import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter import Scrollbar
from tabulate import tabulate
import csv
import openpyxl
import os
import sys

# Global variable to store the loaded data
loaded_data = []
file_path_entry = ""

########################################
##### TO KNOW THE FILE'S SEPARATOR #####
########################################
def detect_separator(file_path):
    with open(file_path, 'r', newline='', encoding='utf-8') as file:
        sample_line = next(file)
        if ',' in sample_line:
            return ','
        elif ';' in sample_line:
            return ';'
        else:
            return None
            
########################################
###### TO READ THE FILE REQUESTED ######
########################################
def read_file(file_path):
    global file_path_entry
    
    if file_path.endswith('.csv'):
        with open(file_path, 'r', newline='', encoding='utf-8') as file:
            try:
                # Determine the delimiter based on the file content
                sample_line = next(file)
            
                if ',' in sample_line:
                    delimiter = ','
                elif ';' in sample_line:
                    delimiter = ';'
                else:
                    messagebox.showinfo("Error", "Unsupported delimiter in the file. Only CSV files with ',' or ';' delimiters are supported.")
                    return

                # Use the determined delimiter when creating the csv.reader
                file.seek(0)  # Reset file pointer to the beginning
                reader = csv.reader(file, delimiter=delimiter)

                for row in reader:
                    yield row
            except StopIteration:
                file_path_entry = ""
                messagebox.showinfo("Error", "The file cannot be read: it's empty.")
                # Handle the error further if needed
                # You can choose to return or exit the program here

    elif file_path.endswith('.xlsx'):
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                yield list(row)
        except Exception as e:
            print(f"Error loading XLSX file: {str(e)}")
    else:
        file_path_entry = ""
        messagebox.showinfo("Error", "Unsupported file format.")



########################################
### TO DISPLAY THE TABLE ON THE GUI ####
########################################
def display_table():
    num_lines_str = num_lines_entry.get()
    if num_lines_str:
        num_lines = int(num_lines_str) + 1  # Get the number of lines from the entry field
    else:
        num_lines = 11  # Display the first 10 lines by default

    table_data = []
    file_path = file_path_entry

    if file_path != "" :
        for i, row in enumerate(read_file(file_path)):
            if i >= num_lines:
                break
            # Add an additional column with the column number to each row
            table_data.append([str(i)] + row)

        table_str = tabulate(table_data, tablefmt="fancy_grid")

        output_text.config(state="normal")
        output_text.delete(1.0, "end")
        output_text.insert("insert", table_str)
        output_text.config(state="disabled")

def show_total_rows():
    file_path = file_path_entry
    if file_path != "" :
        total_rows = sum(1 for _ in read_file(file_path))
        total_rows -= 1
        total_rows_label.config(text=f"Total Rows Available: {total_rows} rows")


#########################################################
### TO SAVE THE TABLE WITH THE NUMBER OF LINES CHOSEN ###
#########################################################
def writeOutput():
    file_path = file_path_entry
    num_lines_str = num_lines_entry.get()

    if num_lines_str:
        num_lines = int(num_lines_str) + 1  # Get the number of lines from the entry field
    else:
        num_lines = None  # None means save all lines

    input_extension = os.path.splitext(file_path)[-1]
    input_filename, _ = os.path.splitext(os.path.basename(file_path))
    output_filename = f"{input_filename}_output{input_extension}"
    
    # Extract the directory path from the loaded file's path
    output_directory = os.path.dirname(file_path)
    output_path = os.path.join(output_directory, output_filename)

    if file_path.endswith('.csv'):
        delimiter = detect_separator(file_path)
        if delimiter is None:
            print("Unsupported delimiter in the file. Only CSV files with ',' or ';' delimiters are supported.")
            return

        # Check if the output file already exists
        counter = 0
        while os.path.exists(output_path):
            counter += 1
            output_filename = f"{input_filename}_output_{counter}{input_extension}"
            output_path = os.path.join(output_directory, output_filename)

        with open(output_path, 'w', newline='', encoding='utf-8') as output_file:
            writer = csv.writer(output_file, delimiter=delimiter)

            for i, row in enumerate(read_file(file_path)):
                if num_lines is not None and i >= num_lines:
                    break
                writer.writerow(row)
        print(f"Table saved to {output_path}")
        messagebox.showinfo("Save Complete", f"Table saved to {output_path}")

    elif file_path.endswith('.xlsx'):
        # Check if the output file already exists
        counter = 0
        while os.path.exists(output_path):
            counter += 1
            output_filename = f"{input_filename}_output_{counter}{input_extension}"
            output_path = os.path.join(output_directory, output_filename)

        output_workbook = openpyxl.Workbook()
        output_sheet = output_workbook.active

        for i, row in enumerate(read_file(file_path)):
            if num_lines is not None and i >= num_lines:
                break
            output_sheet.append(row)

        output_workbook.save(output_path)
        print(f"Table saved to {output_path}")
        messagebox.showinfo("Save Complete", f"Table saved to {output_path}")


########################################
###### TO CHOOSE THE FILE TO LOAD ######
########################################
def open_file_dialog():
    global file_path_entry
    file_path = filedialog.askopenfilename(filetypes=[("All files", "*.*"), ("CSV Files", "*.csv"), ("Excel Files", "*.xlsx")])
    if file_path:
        file_path_entry = file_path
        on_entry_change(None)  # Call on_entry_change when a file is selected
        display_table()  # Display the table when a file is selected

###################################################
### CHANGES TRIGGERED WHEN A NEW FILE IS LOADED ###
###################################################
def on_entry_change(event):
    # This function will be called whenever the entry content changes
    total_rows_label.config(text=f"Total Rows Available: ???? rows")
    num_lines_entry.delete(0, "end")
    output_text.config(state="normal")
    output_text.delete(1.0, "end")

# ***********************************************
# ***********************************************
# **************** PROGRAM'S GUI ****************
# ***********************************************
# ***********************************************

# Initializing the GUI
app = tk.Tk()

# Setting the program's icon
icon_path = os.path.join(sys._MEIPASS, 'icon.ico')
app.iconbitmap(icon_path)

# App title
app.title("CSViewer")
# Set the default window size
app.geometry("800x500")

# Setting the app's frame
frame = ttk.Frame(app)
frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

# Label to choose (and show) the number of available lines
num_available_lines_label = ttk.Label(frame, text="Number of Lines:")
num_available_lines_label.grid(row=1, column=0, padx=5, pady=5)
# Entry to enter the number of lines wished
num_lines_entry = ttk.Entry(frame)
num_lines_entry.grid(row=1, column=1, padx=5, pady=5)

# Button to show the table asked
display_button = ttk.Button(frame, text="Display Table", command=display_table)
display_button.grid(row=1, column=2, padx=5, pady=5)

# Button to save an output of the table, with the number of lines chosen (everything if num_lines_entry is empty)
save_output = ttk.Button(frame, text="Save Table", command=writeOutput)
save_output.grid(row=1, column=3, padx=5, pady=5)

# Output to show the table's data
output_text = tk.Text(frame, state="disabled", wrap="none")
output_text.grid(row=2, column=0, columnspan=4, padx=5, pady=5, sticky="nsew")

# Add vertical scrollbar for output_text
vertical_scrollbar = Scrollbar(frame, orient="vertical", command=output_text.yview)
vertical_scrollbar.grid(row=2, column=4, sticky="ns")
output_text.config(yscrollcommand=vertical_scrollbar.set)

# Add horizontal scrollbar for output_text
horizontal_scrollbar = Scrollbar(frame, orient="horizontal", command=output_text.xview)
horizontal_scrollbar.grid(row=3, column=0, columnspan=4, sticky="ew")
output_text.config(xscrollcommand=horizontal_scrollbar.set)

# Button to show the total number of rows
show_total_rows_button = ttk.Button(frame, text="Show Total Rows", command=show_total_rows)
show_total_rows_button.grid(row=4, column=3, padx=5, pady=5)

# Button to open a file
open_button = ttk.Button(frame, text="Open File", command=open_file_dialog)
open_button.grid(row=4, column=0, padx=5, pady=5)

# Label to display the total number of rows
total_rows_label = ttk.Label(frame, text="Total Rows Available: ???? rows")
total_rows_label.grid(row=4, column=0, columnspan=4, padx=5, pady=5)

# Configure grid for resizing relatively to the frame's dimensions
frame.grid_rowconfigure(2, weight=1)
frame.grid_columnconfigure((0, 1, 2, 3), weight=1)

# Check if the program has been executed from a file; if there is at least one command-line argument
if len(sys.argv) > 1:
    file_path = sys.argv[1]
    file_path_entry = file_path
    print(file_path_entry)
    display_table()

# Loop to make the GUI working
app.mainloop()
